"""Takeoff output: pure document builders + a thin file-writing wrapper.

``build_takeoff_documents`` is pure — ``TakeoffResult`` in, a list of
``(filename, text)`` out — so the CSV/diagnostics shape is unit-testable without
touching the filesystem. ``write_takeoff_export`` is the only I/O: it drops those
documents into a timestamped folder. Mirrors the sibling project's
``build_*_documents`` / ``write_*`` split.

Three documents:
  * ``takeoff_by_system.csv`` — the headline: trusted LF totaled by system.
  * ``takeoff_detail.csv``    — one row per (sheet, style), with provenance and
                                a FLAGGED column for ambiguous styles.
  * ``diagnostics.txt``       — scales, per-sheet notes, errors, flagged styles.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from pathlib import Path

from .models import StyleKey, TakeoffResult


def _fmt_style(k: StyleKey) -> str:
    col = "none" if k.stroke_color is None else ",".join(f"{c:.2f}" for c in k.stroke_color)
    return f"[{col}] w={k.width} {k.dashes}"


def _slug(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", name.strip()).strip("_")
    return s or "takeoff"


def _by_system_csv(result: TakeoffResult) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["System", "Quantity", "Unit", "Sheets", "Styles", "Min_Confidence"])
    by: dict[str, dict] = {}
    for it in result.trusted_items:
        d = by.setdefault(it.system, {"qty": 0.0, "sheets": set(), "styles": set(), "conf": set()})
        d["qty"] += it.quantity
        d["sheets"].add(it.sheet)
        d["styles"].add(it.style_key)
        d["conf"].add(it.confidence)
    order = {"high": 0, "medium": 1, "low": 2}
    for system, d in sorted(by.items(), key=lambda kv: -kv[1]["qty"]):
        worst = max(d["conf"], key=lambda c: order.get(c, 3)) if d["conf"] else ""
        w.writerow([system, f"{d['qty']:.1f}", "LF", len(d["sheets"]), len(d["styles"]), worst])
    return buf.getvalue()


def _detail_csv(result: TakeoffResult) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(
        ["System", "Sheet", "Lineweight", "Quantity", "Unit", "Confidence",
         "Flagged", "Runs", "Scale_pt_per_ft", "Notes"]
    )
    for it in sorted(result.items, key=lambda i: (i.sheet, -i.quantity)):
        w.writerow([
            it.system, it.sheet, _fmt_style(it.style_key), f"{it.quantity:.1f}", it.unit,
            it.confidence, "YES" if not it.trusted else "", it.run_count,
            f"{it.scale_used:g}", (it.reasoning or "").replace("\n", " "),
        ])
    return buf.getvalue()


def _diagnostics_txt(result: TakeoffResult) -> str:
    lines = [
        "=== drawing-takeoff diagnostics ===",
        f"sheets processed: {result.sheet_count}",
        f"takeoff items: {len(result.items)}  (trusted: {len(result.trusted_items)}, "
        f"flagged: {len(result.flagged)})",
        "",
        "TOTALS by system (trusted):",
    ]
    if result.per_system_totals:
        lines += [f"  {sys}: {qty:,.1f} LF" for sys, qty in result.per_system_totals.items()]
    else:
        lines.append("  (none)")
    if result.flagged:
        lines += ["", "FLAGGED for review (not counted):"]
        for it in result.flagged:
            lines.append(
                f"  {it.sheet}  {_fmt_style(it.style_key)}  {it.quantity:,.1f} LF  "
                f"-> {it.system} [conf={it.confidence}, ambiguous={it.ambiguous}]"
                + (f"  — {it.reasoning}" if it.reasoning else "")
            )
    if result.errors:
        lines += ["", "ERRORS:"] + [f"  {e}" for e in result.errors]
    lines += ["", "per-sheet trail:"] + [f"  {d}" for d in result.diagnostics]
    return "\n".join(lines) + "\n"


def build_takeoff_documents(result: TakeoffResult) -> list[tuple[str, str]]:
    """Pure: a takeoff result -> ``[(filename, text), ...]`` (no I/O)."""
    return [
        ("takeoff_by_system.csv", _by_system_csv(result)),
        ("takeoff_detail.csv", _detail_csv(result)),
        ("diagnostics.txt", _diagnostics_txt(result)),
    ]


def write_takeoff_export(
    result: TakeoffResult, out_dir: str | Path = ".", *, project_name: str = "takeoff"
) -> Path:
    """Write the takeoff documents into a ``<slug>_<timestamp>`` folder; return it."""
    folder = Path(out_dir) / f"{_slug(project_name)}_{datetime.now():%Y%m%d_%H%M%S}"
    folder.mkdir(parents=True, exist_ok=True)
    for name, content in build_takeoff_documents(result):
        (folder / name).write_text(content, encoding="utf-8")
    return folder


# ---------------------------------------------------------------------------
# M8: Excel workbook (System x Size summary + per-network detail + review)
# ---------------------------------------------------------------------------
# Nominal-size order for the Summary tab (labels match measure.size_label);
# 'unsized' sorts last. Kept local so export stays decoupled from measure.
_SIZE_ORDER = {
    '1/2"': 0.5, '3/4"': 0.75, '1"': 1.0, '1-1/4"': 1.25, '1-1/2"': 1.5,
    '2"': 2.0, '2-1/2"': 2.5, '3"': 3.0, '4"': 4.0, '6"': 6.0, '8"': 8.0, "unsized": 1e9,
}


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)


def build_takeoff_workbook(tables: dict):
    """Pure: a takeoff -> an openpyxl ``Workbook`` (Summary / Detail / Review tabs).

    ``tables`` is the plain-data shape from :func:`drawing_takeoff.legend.takeoff_tables`
    (``by_system_size``, ``detail`` rows, ``review`` notes), so this writer never
    touches the engine's model types. openpyxl is imported lazily so the rest of
    ``export`` stays import-light.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    bold = Font(bold=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["System", "Size", "Linear Feet"])
    for c in ws[1]:
        c.font = bold
    systems: dict[str, dict[str, float]] = {}
    for (system, size), lf in tables.get("by_system_size", {}).items():
        systems.setdefault(system, {})[size] = lf
    for system in sorted(systems, key=lambda s: -sum(systems[s].values())):
        for size, lf in sorted(systems[system].items(), key=lambda kv: _SIZE_ORDER.get(kv[0], 1e8)):
            ws.append([system, size, round(lf, 1)])
        ws.append([f"{system} — total", "", round(sum(systems[system].values()), 1)])
        ws.cell(ws.max_row, 1).font = bold
        ws.cell(ws.max_row, 3).font = bold
    if not systems:
        ws.append(["(no trusted pipe networks)", "", ""])

    ws2 = wb.create_sheet("Detail")
    detail = tables.get("detail", [])
    # A multi-sheet set tags each row with its sheet (network ids repeat per sheet);
    # the single-sheet CLI export omits it, so the column only appears when needed.
    has_sheet = any("sheet" in r for r in detail)
    header = (["Sheet"] if has_sheet else []) + [
        "Network", "System", "Is Pipe", "Counted", "Confidence", "Ambiguous",
        "Linear Feet", "% Page", "Sizes", "Reasoning",
    ]
    ws2.append(header)
    for c in ws2[1]:
        c.font = bold
    for r in detail:
        row = [
            r["network"], r["system"], "yes" if r["is_pipe"] else "no",
            "yes" if r["counted"] else "no", r["confidence"], "yes" if r["ambiguous"] else "no",
            r["linear_feet"], r["pct_page"], r["sizes"], r["reasoning"],
        ]
        ws2.append(([r.get("sheet", "")] + row) if has_sheet else row)

    ws3 = wb.create_sheet("Review")
    ws3.append(["Review — confirm / not counted"])
    ws3["A1"].font = bold
    for note in tables.get("review", []):
        ws3.append([note])

    for sheet in (ws, ws2, ws3):
        _autosize(sheet)
    return wb


def build_counts_workbook(tables: dict):
    """Pure: a counts takeoff -> an openpyxl ``Workbook`` (Summary / Detail / Review).

    ``tables`` is the plain-data shape from :func:`drawing_takeoff.count.count_tables`
    (``by_component``, ``detail`` rows, ``review`` notes) — aggregated across a
    set by :class:`drawing_takeoff.models.CountsResult.as_tables`. No grand
    total row on purpose: components are heterogeneous (you don't sum WCs and
    sprinkler heads).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    bold = Font(bold=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Component", "Count", "Unit"])
    for c in ws[1]:
        c.font = bold
    by = tables.get("by_component", {})
    for component, n in sorted(by.items(), key=lambda kv: -kv[1]):
        ws.append([component, int(n), "EA"])
    if not by:
        ws.append(["(no trusted countable clusters)", "", ""])

    ws2 = wb.create_sheet("Detail")
    detail = tables.get("detail", [])
    # A multi-sheet set tags each row with its sheet (cluster ids repeat per
    # sheet); the single-sheet CLI export omits it, so the column only appears
    # when needed — same convention as the System×Size Detail tab.
    has_sheet = any("sheet" in r for r in detail)
    header = (["Sheet"] if has_sheet else []) + [
        "Cluster", "Component", "Countable", "Counted", "Confidence", "Ambiguous",
        "Count", "Drawn Size", "Variants", "Reasoning",
    ]
    ws2.append(header)
    for c in ws2[1]:
        c.font = bold
    for r in detail:
        row = [
            r["cluster"], r["component"], "yes" if r["countable"] else "no",
            "yes" if r["counted"] else "no", r["confidence"], "yes" if r["ambiguous"] else "no",
            r["count"], r["size"], r["variants"], r["reasoning"],
        ]
        ws2.append(([r.get("sheet", "")] + row) if has_sheet else row)

    ws3 = wb.create_sheet("Review")
    ws3.append(["Review — confirm / not counted"])
    ws3["A1"].font = bold
    for note in tables.get("review", []):
        ws3.append([note])

    for sheet in (ws, ws2, ws3):
        _autosize(sheet)
    return wb
